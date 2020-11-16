# Gra kółko i krzyżyk

import datetime
import openpyxl
import turtle
from openpyxl import Workbook
from openpyxl.chart import (Reference, Series, BarChart3D, BarChart)
def Start():
    wygranaX=0
    wygranaO=0
    # Funkcja data i czas początkowy gry
    def CzasStart():
        global czasStartu
        czasStartu = datetime.datetime.now()
        print('Start gry:', czasStartu)

    # Funkcja data i czas końcowy gry
    def CzasKoniec():
        global czasKonca
        czasKonca = datetime.datetime.now()
        print('Koniec gry:', czasKonca)

    # Funkcja odmierzająca czas między początkiem a końcem
    def CzasRozgrywki():
        global czasRozgrywki
        czasRozgrywki = czasKonca - czasStartu
        print(str(czasRozgrywki.seconds) + ' sek')

    # Funkcja wygrana X czy O
    def addWin(player):
        nonlocal wygranaX
        nonlocal wygranaO
        if player == 'X':
            wygranaX = wygranaX+1
        else:
            wygranaO = wygranaO+1

    # Funkcja zapisu do Excela
    def Zapiszdane():
        book = openpyxl.load_workbook('//Users//ninakonarzewska//Desktop//WynikiGry.xlsx')
        #sheet = book.get_sheet_by_name('Arkusz1')
        sheet = book['Arkusz1']
        print('Liczba wierszy')
        print(sheet.max_row)
        column = 1
        row=1
        if sheet.max_row == 1:
            sheet.cell(row, column).value = str('Czas Startu')
            sheet.cell(row, column + 1).value = str('Czas Końca')
            sheet.cell(row, column + 2).value = str('Czas Rozgrywki')
            sheet.cell(row, column + 3).value = str('Liczba wygranych X')
            sheet.cell(row, column + 4).value = str('Liczba wygranych 0')
            row = sheet.max_row + 1
            sheet.cell(row, column).value = str(czasStartu)
            sheet.cell(row, column + 1).value = str(czasKonca)
            sheet.cell(row, column + 2).value = str(czasRozgrywki.seconds)
            sheet.cell(row, column + 3).value = wygranaX
            sheet.cell(row, column + 4).value = wygranaO
            # wykres jeżeli pierwszy wpis
            sheet = book.active
            rows = sheet.rows
            columns = sheet.columns
            values = Reference(sheet, min_col=4, min_row=2, max_col=5, max_row=10)
            # labels = Reference(sheet, min_col=3, min_row=2, max_col=5, max_row=2)
            chart = BarChart()
            chart.add_data(values)
            sheet.add_chart(chart, "E15")
        else:
            row = sheet.max_row + 1
            sheet.cell(row, column).value = str(czasStartu)
            sheet.cell(row, column + 1).value = str(czasKonca)
            sheet.cell(row, column + 2).value = str(czasRozgrywki.seconds)
            sheet.cell(row, column + 3).value = wygranaX
            sheet.cell(row, column + 4).value = wygranaO



        book.save('//Users//ninakonarzewska//Desktop//WynikiGry.xlsx')
        print('saved')

    # Funkcja zawierająca całą funkcjonalność gry
    def restart():
        print('Witam w grze Kółko i krzyżyk!')
        Start = CzasStart()

        # Opis planszy według pozycji w polu
        theBoard = {'top-L': ' ', 'top-M': ' ', 'top-R': ' ',
                    'mid-L': ' ', 'mid-M': ' ', 'mid-R': ' ',
                    'low-L': ' ', 'low-M': ' ', 'low-R': ' '}

        # Funkcja rysowania planszy gry
        def printBoard(board):
            print(board['top-L'] + '|' + board['top-M'] + '|' + board['top-R'])
            print('-+-+-')
            print(board['mid-L'] + '|' + board['mid-M'] + '|' + board['mid-R'])
            print('-+-+-')
            print(board['low-L'] + '|' + board['low-M'] + '|' + board['low-R'])

        turn = 'X'

        # Funkcja sprawdzenia miejsc wygranych w poziomie, w pionie i na skos planszy
        def winner(board):
            won = False
            if (board['top-L'] == board['top-M'] == board['top-R'] != ' '):
                won = True
            elif (board['mid-L'] == board['mid-M'] == board['mid-R'] != ' '):
                won = True
            elif (board['low-L'] == board['low-M'] == board['low-R'] != ' '):
                won = True
            elif (board['top-L'] == board['mid-L'] == board['low-L'] != ' '):
                won = True
            elif (board['top-M'] == board['mid-M'] == board['low-M'] != ' '):
                won = True
            elif (board['top-R'] == board['mid-M'] == board['low-R'] != ' '):
                won = True
            elif (board['top-L'] == board['mid-M'] == board['low-R'] != ' '):
                won = True
            elif (board['top-R'] == board['mid-M'] == board['low-L'] != ' '):
                won = True
            return won

        for i in range(9):
            printBoard(theBoard)
            while True:
                try:
                    print('Kolej na  ' + turn + '. Ruch, na którą cześć planszy?')
                    move = str(input())
                    theBoard[move] = turn
                    break
                except ValueError:
                    print('Ups! To nie był odpowiedni ruch')
            won = winner(theBoard)
            if won == 1:
                print(turn + ' wygrywa grę!')
                addWin(turn)
                break
            else:
                print('Nie ma wygranej, zagraj jeszcze raz')
            if turn == 'X':
                turn = '0'
            else:
                turn = 'X'
        printBoard(theBoard)

        # Ponowna gra lub zakończenie
        answer = input("Czy chcesz zagrać jeszcze raz?Tak lub Nie\n")
        if answer == 'Tak':
            print('Zagrajmy jeszcze raz!')
            Koniec = CzasKoniec()
            Rozgrywka = CzasRozgrywki()
            #Zapiszdane()
            restart()
        else:
            print('Koniec gry na dzisiaj')
            Koniec = CzasKoniec()
            Rozgrywka = CzasRozgrywki()
            Zapiszdane()
            turtle.circle(100)
            img = Image.open('X-image.jpg')
            img.show()

    restart()
Start()
